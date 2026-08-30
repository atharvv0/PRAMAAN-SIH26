from __future__ import annotations

from pathlib import Path


def extract_document_text(path: str) -> str:
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(str(p))

    suffix = p.suffix.lower()

    if suffix == ".pdf":
        import fitz

        with fitz.open(p) as doc:
            return "\n".join(page.get_text("text") for page in doc)

    if suffix == ".docx":
        from docx import Document

        doc = Document(p)
        return "\n".join(paragraph.text for paragraph in doc.paragraphs)

    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(p, read_only=True, data_only=True)
        lines: list[str] = []
        for ws in wb.worksheets:
            lines.append(f"[SHEET: {ws.title}]")
            for row in ws.iter_rows(values_only=True):
                values = [str(value) for value in row if value is not None]
                if values:
                    lines.append(" | ".join(values))
        wb.close()
        return "\n".join(lines)

    return p.read_text(encoding="utf-8")
